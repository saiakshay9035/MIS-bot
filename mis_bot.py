import streamlit as st
import pandas as pd
import io
import datetime

def main():
    st.title("🤖 MIS Support Bot")
    
    # Greeting
    st.write("Hi team, how are you doing today? Please upload your raw Excel file to proceed.")
    
    # File uploader
    uploaded_file = st.file_uploader(
        "Choose your file", 
        type=['xlsx', 'csv'],
        help="Upload Excel (.xlsx) or CSV file"
    )
    
    if uploaded_file is not None:
        # Load data
        try:
            if uploaded_file.name.endswith('.xlsx'):
                df = pd.read_excel(uploaded_file)
            else:
                df = pd.read_csv(uploaded_file)
            
            st.success(f"✅ File uploaded successfully! ({len(df)} rows)")
            
            # Show data preview
            with st.expander("📊 Data Preview"):
                st.dataframe(df.head())
            
            # MIS Type Selection
            st.subheader("Select MIS Type:")
            
            mis_options = [
                "Client MIS",
                "Open Ticket MIS", 
                "Request Ticket Open MIS",
                "Request Ticket Closed MIS",
                "Bug Ticket Closed MIS",
                "Jagan's MIS",
                "Recurring Issues MIS"
            ]
            
            selected_mis = st.radio("Choose MIS type:", mis_options)
            
            if st.button("Generate MIS", type="primary"):
                # Process MIS (placeholder for your backend logic)
                processed_df = process_mis(df, selected_mis)
                
                st.success(f"✅ {selected_mis} generated successfully!")
                
                # Display results
                st.subheader("📈 MIS Results:")
                if selected_mis == "Client MIS" and isinstance(processed_df, dict) and len(processed_df) > 1:
                    st.write(f"**Generated MIS for {len(processed_df)} programs:**")
                    for program_name, program_data in processed_df.items():
                        with st.expander(f"📊 {program_name} MIS"):
                            if isinstance(program_data, dict):
                                st.write("**MIS Report:**")
                                st.dataframe(program_data['mis_report'])
                                st.write("**Open Tickets:**")
                                st.dataframe(program_data['open_data'])
                                st.write("**Closed Tickets:**")
                                st.dataframe(program_data['closed_data'])
                                st.write("**Request Tickets:**")
                                st.dataframe(program_data['request_data'])
                            else:
                                st.dataframe(program_data)
                elif isinstance(processed_df, dict) and 'raw_data' in processed_df:
                    st.write("**Raw Data:**")
                    st.dataframe(processed_df['raw_data'])
                    st.write("**MIS Summary:**")
                    st.dataframe(processed_df['mis_summary'])
                else:
                    st.dataframe(processed_df)
                
                # Download button
                excel_buffer = io.BytesIO()
                
                if selected_mis == "Request Ticket Open MIS" and isinstance(processed_df, dict):
                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        processed_df['raw_data'].to_excel(writer, index=False, sheet_name='Request Open Ticket')
                        processed_df['mis_summary'].to_excel(writer, index=False, sheet_name='MIS')
                    
                    st.download_button(
                        label="📥 Download MIS as Excel",
                        data=excel_buffer.getvalue(),
                        file_name=f"Request_open_ticket_{datetime.datetime.now().strftime('%d-%b')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                elif selected_mis == "Client MIS":
                    # Handle multiple program files
                    if isinstance(processed_df, dict) and len(processed_df) > 1:
                        # Create a zip file with multiple Excel files
                        import zipfile
                        zip_buffer = io.BytesIO()
                        
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                            for program_name, program_data in processed_df.items():
                                excel_buffer = io.BytesIO()
                                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                    # Write MIS report
                                    program_data['mis_report'].to_excel(writer, index=False, sheet_name='Client_MIS', header=False)
                                    # Write separate sheets for different ticket types
                                    program_data['open_data'].to_excel(writer, index=False, sheet_name='Open_Tickets')
                                    program_data['closed_data'].to_excel(writer, index=False, sheet_name='Closed_Tickets')
                                    program_data['request_data'].to_excel(writer, index=False, sheet_name='Request_Tickets')
                                
                                safe_program_name = program_name.replace('/', '_').replace('\\', '_')
                                zip_file.writestr(f"{safe_program_name}_client_mis_{datetime.datetime.now().strftime('%d-%b')}.xlsx", excel_buffer.getvalue())
                        
                        st.download_button(
                            label="📥 Download All Program MIS as ZIP",
                            data=zip_buffer.getvalue(),
                            file_name=f"client_mis_all_programs_{datetime.datetime.now().strftime('%d-%b')}.zip",
                            mime="application/zip"
                        )
                    else:
                        # Single program or error case
                        if isinstance(processed_df, dict):
                            program_data = list(processed_df.values())[0]
                            if isinstance(program_data, dict):
                                excel_buffer = io.BytesIO()
                                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                    # Write MIS report
                                    program_data['mis_report'].to_excel(writer, index=False, sheet_name='Client_MIS', header=False)
                                    # Write separate sheets for different ticket types
                                    program_data['open_data'].to_excel(writer, index=False, sheet_name='Open_Tickets')
                                    program_data['closed_data'].to_excel(writer, index=False, sheet_name='Closed_Tickets')
                                    program_data['request_data'].to_excel(writer, index=False, sheet_name='Request_Tickets')
                            else:
                                excel_buffer = io.BytesIO()
                                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                    program_data.to_excel(writer, index=False, sheet_name='Client_MIS', header=False)
                        else:
                            excel_buffer = io.BytesIO()
                            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                processed_df.to_excel(writer, index=False, sheet_name='Client_MIS', header=False)
                        
                        st.download_button(
                            label="📥 Download Client MIS as Excel",
                            data=excel_buffer.getvalue(),
                            file_name=f"client_mis_{datetime.datetime.now().strftime('%d-%b')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                
                elif selected_mis in ["Open Ticket MIS", "Bug Ticket Closed MIS", "Jagan's MIS", "Recurring Issues MIS"]:
                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        processed_df.to_excel(writer, index=False, sheet_name=selected_mis.replace(' ', '_'), header=False)
                        
                        # Add red highlighting for crossed SLA tickets if applicable
                        if selected_mis in ["Open Ticket MIS", "Jagan's MIS"]:
                            from openpyxl.styles import PatternFill
                            import re
                            worksheet = writer.sheets[selected_mis.replace(' ', '_')]
                            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                            
                            # Highlight rows with any Crossed SLA percentage > 0%
                            for row in worksheet.iter_rows():
                                should_highlight = False
                                for cell in row:
                                    if cell.value:
                                        cell_str = str(cell.value)
                                        # Check for any percentage > 0% in Crossed SLA column
                                        if re.match(r'^([1-9]\d*|[1-9])%$', cell_str.strip()):
                                            # Check if this cell is in a Crossed SLA% column
                                            col_header = worksheet.cell(row=1, column=cell.column).value
                                            if col_header and 'Crossed SLA%' in str(col_header):
                                                should_highlight = True
                                                break
                                        # Also highlight rows containing "Crossed SLA" text
                                        elif 'Crossed SLA' in cell_str:
                                            should_highlight = True
                                            break
                                
                                if should_highlight:
                                    for cell_in_row in row:
                                        cell_in_row.fill = red_fill
                    
                    st.download_button(
                        label="📥 Download MIS as Excel",
                        data=excel_buffer.getvalue(),
                        file_name=f"{selected_mis.replace(' ', '_').lower()}_{datetime.datetime.now().strftime('%d-%b')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                else:
                    # For other MIS types, provide CSV download
                    csv_buffer = io.StringIO()
                    processed_df.to_csv(csv_buffer, index=False)
                    
                    st.download_button(
                        label="📥 Download MIS as CSV",
                        data=csv_buffer.getvalue(),
                        file_name=f"{selected_mis.replace(' ', '_').lower()}.csv",
                        mime="text/csv"
                    )
                
        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")

def process_mis(df, mis_type):
    """
    Process MIS based on the selected type
    """
    if mis_type == "Open Ticket MIS":
        return process_open_ticket_mis(df)
    elif mis_type == "Client MIS":
        return process_client_mis(df)
    elif mis_type == "Request Ticket Open MIS":
        return process_request_ticket_open_mis(df)
    elif mis_type == "Request Ticket Closed MIS":
        return process_request_ticket_closed_mis(df)
    elif mis_type == "Bug Ticket Closed MIS":
        return process_bug_ticket_closed_mis(df)
    elif mis_type == "Jagan's MIS":
        return process_jagan_mis(df)
    elif mis_type == "Recurring Issues MIS":
        return process_recurring_issues_mis(df)
    
    return df

def process_client_mis(df):
    """Process Client MIS - Generate program wise MIS with 3 sections each"""
    if 'Program Name' not in df.columns:
        return pd.DataFrame({'Error': ['Program Name column not found']})
    
    # Filter only client tickets (exclude internal tickets)
    if 'Ticket Group' in df.columns:
        client_df = df[df['Ticket Group'].str.lower().str.contains('client', na=False)].copy()
    else:
        return pd.DataFrame({'Error': ['Ticket Group column not found']})
    
    if client_df.empty:
        return pd.DataFrame({'Error': ['No client tickets found']})
    
    # Get unique programs from client tickets only
    programs = client_df['Program Name'].unique()
    program_reports = {}
    
    for program in programs:
        # Filter data for this program only (already filtered for client tickets)
        program_df = client_df[client_df['Program Name'] == program].copy()
        
        if program_df.empty:
            continue
            
        # Calculate SLA status
        if 'Is Overdue' in program_df.columns:
            program_df['SLA_Status'] = program_df['Is Overdue'].apply(
                lambda x: 'Crossed SLA' if x == True else 'Within SLA'
            )
        else:
            program_df['SLA_Status'] = 'Within SLA'
        
        # Generate 3 sections for this program
        final_report = []
        
        # 1. Closed Tickets Section
        final_report.append([f'{program} - Closed Tickets:'])
        final_report.append([''])
        closed_tickets = program_df[program_df['Status (Ticket)'] == 'Closed']
        
        if not closed_tickets.empty:
            closed_report = generate_client_closed_report(closed_tickets, program)
            final_report.extend(closed_report.values.tolist())
        else:
            final_report.append(['Client Name', 'Closed Tickets Within SLA', 'Closed Tickets Crossed SLA', 'Total Closed Tickets', 'Within SLA%', 'Crossed SLA%'])
            final_report.append([program, 0, 0, 0, '0%', '0%'])
        
        final_report.append([''])
        final_report.append([''])
        
        # 2. Open Tickets Section
        final_report.append([f'{program} - Open Tickets:'])
        final_report.append([''])

        # Ensure consistent casing and trim spaces
        program_df['Status (Ticket)'] = program_df['Status (Ticket)'].str.strip().str.lower()

        # Define open statuses (lowercase)
        open_statuses = [
            'assigned to engineer!',
            'reopened',
            'waiting information from user - 1',
            'waiting information from user - 2',
            'waiting information from user - 3'
        ]
        open_tickets = program_df[program_df['Status (Ticket)'].isin(open_statuses)].copy()

        # Waiting statuses (lowercase)
        waiting_statuses = [
            'waiting information from user - 1',
            'waiting information from user - 2',
            'waiting information from user - 3'
        ]

        # Apply exclude rule if Classifications exists
        if 'Classifications' in open_tickets.columns:
            open_tickets['Classifications'] = open_tickets['Classifications'].str.strip().str.lower()
            exclude_condition = (
                open_tickets['Status (Ticket)'].isin(waiting_statuses) &
                (open_tickets['Classifications'] == 'request open')
            )
            open_tickets = open_tickets[~exclude_condition]

        if not open_tickets.empty:
            open_report = generate_client_open_report(open_tickets, program)
            final_report.extend(open_report.values.tolist())
        else:
            final_report.append(['Client Name', 'Open tickets within SLA', 'Open Tickets Crossed SLA', 'Total Open Tickets', 'Within SLA%', 'Crossed SLA%'])
            final_report.append([program, 0, 0, 0, '0%', '0%'])

        final_report.append([''])
        final_report.append([''])

        
        # 3. Request Tickets Section
        final_report.append([f'{program} - Request Tickets:'])
        final_report.append([''])
        request_report = generate_client_request_report(program_df, program)
        final_report.extend(request_report.values.tolist())
        
        # Prepare raw data with specified columns
        raw_data_columns = [
            'Ticket Id', 'Status (Ticket)', 'Created Time (Ticket)', 'Due Date', 
            'Email (Contact)', 'Priority (Ticket)', 'Program Name', 'Crossed Due Date', 
            'Request Sub Category', 'Contact name'
        ]
        
        # Map column names if needed
        column_mapping = {
            'Created Tim': 'Created Time (Ticket)',
            'Account Name': 'Contact name'
        }
        
        program_df_mapped = program_df.copy()
        for old_name, new_name in column_mapping.items():
            if old_name in program_df_mapped.columns and new_name not in program_df_mapped.columns:
                program_df_mapped = program_df_mapped.rename(columns={old_name: new_name})
        
        # Select only available columns from the specified list
        available_raw_columns = [col for col in raw_data_columns if col in program_df_mapped.columns]
        base_raw_data = program_df_mapped[available_raw_columns].copy()
        
        # Use Program Name as Client Name in raw data
        if 'Program Name' in base_raw_data.columns:
            base_raw_data = base_raw_data.rename(columns={'Program Name': 'Client Name'})
        
        # Separate data by ticket type
        closed_data = base_raw_data[base_raw_data['Status (Ticket)'] == 'Closed'].copy()
        
        open_statuses = ['Assigned to Engineer!', 'Reopened', 'Waiting Information From user - 1', 'Waiting Information From user - 2', 'Waiting Information From user - 3']
        open_data = base_raw_data[base_raw_data['Status (Ticket)'].isin(open_statuses)].copy()
        
        # Filter request tickets based on Classifications column
        if 'Classifications' in program_df_mapped.columns:
            request_filter = program_df_mapped['Classifications'].str.lower().str.contains('request', na=False)
            request_data = base_raw_data[request_filter].copy()
        else:
            # If no Classifications column, assume all are request tickets
            request_data = base_raw_data.copy()
        
        program_reports[program] = {
            'mis_report': pd.DataFrame(final_report),
            'open_data': open_data,
            'closed_data': closed_data,
            'request_data': request_data
        }
    
    return program_reports

def process_request_ticket_closed_mis(df):
    """Process Request Ticket Closed MIS"""
    if 'Status (Ticket)' not in df.columns:
        return pd.DataFrame({'Error': ['Status (Ticket) column not found']})
    
    closed_tickets = df[df['Status (Ticket)'] == 'Closed'].copy()
    
    if closed_tickets.empty:
        return pd.DataFrame({'Message': ['No closed request tickets found']})
    
    summary = closed_tickets.groupby(['Select Engineer', 'Priority (Ticket)']).size().unstack(fill_value=0)
    summary['Total'] = summary.sum(axis=1)
    return summary.reset_index()

def process_bug_ticket_closed_mis(df):
    """Process Bug Ticket Closed MIS - similar to Open Ticket MIS but for closed bug tickets"""
    # Debug: show available columns
    available_cols = list(df.columns)
    
    # Check if we have the basic columns needed
    if 'Status (Ticket)' not in df.columns:
        return pd.DataFrame({'Error': [f'Status (Ticket) column not found. Available columns: {available_cols[:10]}...']})
    
    # Filter only by closed status
    closed_bug_tickets = df[
        df['Status (Ticket)'].isin(['Closed', 'Closed due to lack of information'])
    ].copy()
    
    if closed_bug_tickets.empty:
        return pd.DataFrame({'Error': ['No closed bug tickets found']})
    
    # Calculate SLA status (check if Is Overdue column exists)
    if 'Is Overdue' in closed_bug_tickets.columns:
        closed_bug_tickets['SLA_Status'] = closed_bug_tickets['Is Overdue'].apply(
            lambda x: 'Crossed SLA' if x == True else 'Within SLA'
        )
    else:
        # If no Is Overdue column, assume all are within SLA
        closed_bug_tickets['SLA_Status'] = 'Within SLA'
    
    # Generate all three reports
    module_lead_report = generate_bug_module_lead_report(closed_bug_tickets)
    client_report = generate_bug_client_report(closed_bug_tickets)
    engineer_report = generate_bug_engineer_report(closed_bug_tickets)
    
    # Create structured report with proper headers
    final_report = []
    
    # Module Lead Report
    final_report.append(['MODULE LEAD WISE REPORT'])
    final_report.extend(module_lead_report.values.tolist())
    final_report.append([''])
    
    # Client Report  
    final_report.append(['CLIENT WISE REPORT'])
    final_report.extend(client_report.values.tolist())
    final_report.append([''])
    
    # Engineer Report
    final_report.append(['ENGINEER WISE REPORT'])
    final_report.extend(engineer_report.values.tolist())
    
    return pd.DataFrame(final_report)

def generate_bug_module_lead_report(closed_bug_tickets):
    """Generate Module Lead wise report for closed bug tickets"""
    report = closed_bug_tickets.groupby(['Module Lead', 'SLA_Status']).size().unstack(fill_value=0)
    
    result = []
    result.append(['Module Lead', 'Closed Bug Within SLA', 'Closed Bug Crossed SLA', 'Total Closed Bugs', 'Within SLA%', 'Crossed SLA%'])
    
    total_within = 0
    total_crossed = 0
    
    for module_lead in report.index:
        within_sla = report.loc[module_lead, 'Within SLA'] if 'Within SLA' in report.columns else 0
        crossed_sla = report.loc[module_lead, 'Crossed SLA'] if 'Crossed SLA' in report.columns else 0
        total = within_sla + crossed_sla
        
        within_pct = f"{int(within_sla * 100 / total)}%" if total > 0 else "0%"
        crossed_pct = f"{int(crossed_sla * 100 / total)}%" if total > 0 else "0%"
        
        result.append([module_lead, within_sla, crossed_sla, total, within_pct, crossed_pct])
        
        total_within += within_sla
        total_crossed += crossed_sla
    
    # Grand total
    grand_total = total_within + total_crossed
    within_pct = f"{int(total_within * 100 / grand_total)}%" if grand_total > 0 else "0%"
    crossed_pct = f"{int(total_crossed * 100 / grand_total)}%" if grand_total > 0 else "0%"
    result.append(['Grand Total', total_within, total_crossed, grand_total, within_pct, crossed_pct])
    
    return pd.DataFrame(result[1:], columns=result[0])

def generate_bug_client_report(closed_bug_tickets):
    """Generate Client wise report for closed bug tickets"""
    # Use Program Name as Client Name
    report = closed_bug_tickets.groupby(['Program Name', 'SLA_Status']).size().unstack(fill_value=0)
    
    result = []
    result.append(['Client Name', 'Closed Bug Within SLA', 'Closed Bug Crossed SLA', 'Total Closed Bugs', 'Within SLA%', 'Crossed SLA%'])
    
    total_within = 0
    total_crossed = 0
    
    for client in report.index:
        within_sla = report.loc[client, 'Within SLA'] if 'Within SLA' in report.columns else 0
        crossed_sla = report.loc[client, 'Crossed SLA'] if 'Crossed SLA' in report.columns else 0
        total = within_sla + crossed_sla
        
        within_pct = f"{int(within_sla * 100 / total)}%" if total > 0 else "0%"
        crossed_pct = f"{int(crossed_sla * 100 / total)}%" if total > 0 else "0%"
        
        result.append([client, within_sla, crossed_sla, total, within_pct, crossed_pct])
        
        total_within += within_sla
        total_crossed += crossed_sla
    
    # Grand total
    grand_total = total_within + total_crossed
    within_pct = f"{int(total_within * 100 / grand_total)}%" if grand_total > 0 else "0%"
    crossed_pct = f"{int(total_crossed * 100 / grand_total)}%" if grand_total > 0 else "0%"
    result.append(['Grand Total', total_within, total_crossed, grand_total, within_pct, crossed_pct])
    
    return pd.DataFrame(result[1:], columns=result[0])

def generate_bug_engineer_report(closed_bug_tickets):
    """Generate Engineer wise report for closed bug tickets"""
    report = closed_bug_tickets.groupby(['Select Engineer', 'SLA_Status']).size().unstack(fill_value=0)
    
    result = []
    result.append(['Engineer', 'Closed Bug Within SLA', 'Closed Bug Crossed SLA', 'Total Closed Bugs', 'Within SLA%', 'Crossed SLA%'])
    
    total_within = 0
    total_crossed = 0
    
    for engineer in report.index:
        within_sla = report.loc[engineer, 'Within SLA'] if 'Within SLA' in report.columns else 0
        crossed_sla = report.loc[engineer, 'Crossed SLA'] if 'Crossed SLA' in report.columns else 0
        total = within_sla + crossed_sla
        
        within_pct = f"{int(within_sla * 100 / total)}%" if total > 0 else "0%"
        crossed_pct = f"{int(crossed_sla * 100 / total)}%" if total > 0 else "0%"
        
        result.append([engineer, within_sla, crossed_sla, total, within_pct, crossed_pct])
        
        total_within += within_sla
        total_crossed += crossed_sla
    
    # Grand total
    grand_total = total_within + total_crossed
    within_pct = f"{int(total_within * 100 / grand_total)}%" if grand_total > 0 else "0%"
    crossed_pct = f"{int(total_crossed * 100 / grand_total)}%" if grand_total > 0 else "0%"
    result.append(['Grand Total', total_within, total_crossed, grand_total, within_pct, crossed_pct])
    
    return pd.DataFrame(result[1:], columns=result[0])

def process_jagan_mis(df):
    """Process Jagan's MIS with 4 specific sections"""
    import datetime
    
    # Check required columns
    if 'Status (Ticket)' not in df.columns:
        return pd.DataFrame({'Error': ['Status (Ticket) column not found']})
    
    # Ensure consistent casing and trim spaces
    df['Status (Ticket)'] = df['Status (Ticket)'].str.strip().str.lower()

    # Define open statuses
    open_statuses = [
        'assigned to engineer!',
        'reopened',
        'waiting information from user - 1',
        'waiting information from user - 2',
        'waiting information from user - 3'
    ]
    open_tickets = df[df['Status (Ticket)'].isin(open_statuses)].copy()

    # Waiting statuses
    waiting_statuses = [
        'waiting information from user - 1',
        'waiting information from user - 2',
        'waiting information from user - 3'
    ]

    # Apply exclude rule
    if 'Classifications' in open_tickets.columns:
        open_tickets['Classifications'] = open_tickets['Classifications'].str.strip().str.lower()
        exclude_condition = (
            open_tickets['Status (Ticket)'].isin(waiting_statuses) &
            (open_tickets['Classifications'] == 'request open')
        )
        open_tickets = open_tickets[~exclude_condition]

    if open_tickets.empty:
        return pd.DataFrame({'Error': ['No open tickets found']})
    
    # Continue with the rest of your Jagan MIS logic below...


    
    # Calculate days from creation
    today_date = datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Calculate SLA status based on Gitlab due date
    def calculate_sla_status_jagan(row):
        if 'Gitlab Due date' in open_tickets.columns and pd.notna(row.get('Gitlab Due date')):
            try:
                gitlab_due = pd.to_datetime(row['Gitlab Due date'], errors='coerce')
                if pd.notna(gitlab_due):
                    return 'Crossed SLA' if today_date.date() > gitlab_due.date() else 'Within SLA'
            except:
                pass
        # Fallback to Is Overdue if Gitlab due date not available
        if 'Is Overdue' in open_tickets.columns:
            return 'Crossed SLA' if row.get('Is Overdue') == True else 'Within SLA'
        return 'Within SLA'
    
    open_tickets['SLA_Status'] = open_tickets.apply(calculate_sla_status_jagan, axis=1)
    
    # Sort by creation date (ascending) to show oldest tickets first
    if 'Created Time (Ticket)' in open_tickets.columns:
        open_tickets['Created_Date_Sort'] = pd.to_datetime(open_tickets['Created Time (Ticket)'], errors='coerce')
        open_tickets = open_tickets.sort_values('Created_Date_Sort', ascending=True)
        open_tickets = open_tickets.drop('Created_Date_Sort', axis=1)
    
    def calculate_days_diff(created_date):
        try:
            if pd.isna(created_date):
                return 0
            created_dt = pd.to_datetime(created_date, errors='coerce')
            if pd.isna(created_dt):
                return 0
            return max(0, (today_date - created_dt).days)
        except:
            return 0
    
    created_col = 'Created Time (Ticket)'
    if created_col in open_tickets.columns:
        open_tickets['Days_Crossed'] = open_tickets[created_col].apply(calculate_days_diff)
    else:
        open_tickets['Days_Crossed'] = 0
    
    final_report = []
    
    # 1. Department wise SLA status for open tickets
    final_report.append(['DEPARTMENT WISE SLA STATUS - OPEN TICKETS'])
    dept_report = open_tickets.groupby(['Department Name', 'SLA_Status']).size().unstack(fill_value=0)
    
    result = []
    result.append(['Department Name', 'Within SLA', 'Crossed SLA', 'Grand Total', 'Within SLA%', 'Crossed SLA%'])
    
    total_within = 0
    total_crossed = 0
    data_rows = []
    
    for dept in dept_report.index:
        within_sla = dept_report.loc[dept, 'Within SLA'] if 'Within SLA' in dept_report.columns else 0
        crossed_sla = dept_report.loc[dept, 'Crossed SLA'] if 'Crossed SLA' in dept_report.columns else 0
        total = within_sla + crossed_sla
        
        within_pct_num = round(within_sla * 100 / total) if total > 0 else 0
        crossed_pct_num = round(crossed_sla * 100 / total) if total > 0 else 0
        # Ensure percentages add up to 100%
        if total > 0 and within_pct_num + crossed_pct_num != 100:
            within_pct_num = 100 - crossed_pct_num
        within_pct = f"{within_pct_num}%"
        crossed_pct = f"{crossed_pct_num}%"
        
        data_rows.append([dept, within_sla, crossed_sla, total, within_pct, crossed_pct, crossed_pct_num])
        
        total_within += within_sla
        total_crossed += crossed_sla
    
    # Sort by crossed SLA percentage (descending)
    data_rows.sort(key=lambda x: x[6], reverse=True)
    
    # Add sorted rows (without the sort key)
    for row in data_rows:
        result.append(row[:6])
    
    # Grand total
    grand_total = total_within + total_crossed
    within_pct_num = round(total_within * 100 / grand_total) if grand_total > 0 else 0
    crossed_pct_num = round(total_crossed * 100 / grand_total) if grand_total > 0 else 0
    if grand_total > 0 and within_pct_num + crossed_pct_num != 100:
        within_pct_num = 100 - crossed_pct_num
    within_pct = f"{within_pct_num}%"
    crossed_pct = f"{crossed_pct_num}%"
    result.append(['Grand Total', total_within, total_crossed, grand_total, within_pct, crossed_pct])
    
    final_report.extend(result)
    final_report.append([''])
    
    # 2. Tickets that crossed SLA with GitLab links
    final_report.append(['TICKETS CROSSED SLA WITH GITLAB LINKS'])
    crossed_sla_tickets = open_tickets[open_tickets['SLA_Status'] == 'Crossed SLA']
    
    if not crossed_sla_tickets.empty:
        header = ['Gitlab Link', 'Select Engineer', 'Program Name', 'Department Name']
        final_report.append(header)
        
        for _, ticket in crossed_sla_tickets.iterrows():
            row = [
                ticket.get('Gitlab Link', ''),
                ticket.get('Select Engineer', ''),
                ticket.get('Program Name', ''),
                ticket.get('Department Name', '')
            ]
            final_report.append(row)
    else:
        final_report.append(['No tickets crossed SLA'])
    
    final_report.append([''])
    
    # 3. Tickets that will cross due date today
    final_report.append(['TICKETS WILL CROSS DUE DATE TODAY'])
    
    # Check if Gitlab Due date column exists and filter tickets due today
    if 'Gitlab Due date' in open_tickets.columns:
        today_str = today_date.strftime('%Y-%m-%d')
        # Convert GitLab due date to datetime and compare with today
        due_today = open_tickets[
            pd.to_datetime(open_tickets['Gitlab Due date'], errors='coerce').dt.date == today_date.date()
        ]
        
        if not due_today.empty:
            header = ['Gitlab Link', 'Select Engineer', 'Program Name', 'Department Name']
            final_report.append(header)
            
            for _, ticket in due_today.iterrows():
                row = [
                    ticket.get('Gitlab Link', ''),
                    ticket.get('Select Engineer', ''),
                    ticket.get('Program Name', ''),
                    ticket.get('Department Name', '')
                ]
                final_report.append(row)
        else:
            final_report.append(['No tickets due today'])
    else:
        final_report.append(['Gitlab Due date column not found'])
    
    final_report.append([''])
    
    # 4. Number of days crossed - Open tickets summary
    final_report.append(['NUMBER OF DAYS CROSSED - OPEN TICKETS'])
    
    # Group tickets by actual days crossed
    days_summary = open_tickets['Days_Crossed'].value_counts().sort_index()
    
    # Create header with actual days that have tickets
    day_cols = sorted([day for day in days_summary.index if days_summary[day] > 0])
    header = ['# of Days'] + [str(day) for day in day_cols] + ['Grand Total']
    final_report.append(header)
    
    # Add count row
    counts = [days_summary.get(day, 0) for day in day_cols]
    total = sum(counts)
    row = ['Count'] + counts + [total]
    final_report.append(row)
    final_report.append([''])
    
    # Add existing Open Ticket MIS reports
    # 5. Module Lead wise report
    final_report.append(['MODULE LEAD WISE REPORT'])
    module_lead_report = generate_module_lead_report(open_tickets)
    final_report.extend(module_lead_report.values.tolist())
    final_report.append([''])
    
    # 6. Client wise report
    final_report.append(['CLIENT WISE REPORT'])
    client_report = generate_client_report(open_tickets)
    final_report.extend(client_report.values.tolist())
    final_report.append([''])
    
    # 7. Engineer wise report
    final_report.append(['ENGINEER WISE REPORT'])
    engineer_report = generate_engineer_report(open_tickets)
    final_report.extend(engineer_report.values.tolist())
    final_report.append([''])
    
    # 8. Product/PS wise report
    final_report.append(['PRODUCT/PS WISE REPORT'])
    if 'Product OR PS Ticket' in open_tickets.columns:
        ps_report = open_tickets.groupby(['Product OR PS Ticket', 'SLA_Status']).size().unstack(fill_value=0)
        
        result = []
        result.append(['Product OR PS Ticket', 'Within SLA', 'Crossed SLA', 'Grand Total', 'Within SLA%', 'Crossed SLA%'])
        
        total_within = 0
        total_crossed = 0
        data_rows = []
        
        for ps_type in ps_report.index:
            within_sla = ps_report.loc[ps_type, 'Within SLA'] if 'Within SLA' in ps_report.columns else 0
            crossed_sla = ps_report.loc[ps_type, 'Crossed SLA'] if 'Crossed SLA' in ps_report.columns else 0
            total = within_sla + crossed_sla
            
            within_pct_num = round(within_sla * 100 / total) if total > 0 else 0
            crossed_pct_num = round(crossed_sla * 100 / total) if total > 0 else 0
            # Ensure percentages add up to 100%
            if total > 0 and within_pct_num + crossed_pct_num != 100:
                within_pct_num = 100 - crossed_pct_num
            within_pct = f"{within_pct_num}%"
            crossed_pct = f"{crossed_pct_num}%"
            
            data_rows.append([ps_type, within_sla, crossed_sla, total, within_pct, crossed_pct, crossed_pct_num])
            
            total_within += within_sla
            total_crossed += crossed_sla
        
        # Sort by crossed SLA percentage (descending)
        data_rows.sort(key=lambda x: x[6], reverse=True)
        
        # Add sorted rows (without the sort key)
        for row in data_rows:
            result.append(row[:6])
        
        # Grand total
        grand_total = total_within + total_crossed
        within_pct_num = round(total_within * 100 / grand_total) if grand_total > 0 else 0
        crossed_pct_num = round(total_crossed * 100 / grand_total) if grand_total > 0 else 0
        if grand_total > 0 and within_pct_num + crossed_pct_num != 100:
            within_pct_num = 100 - crossed_pct_num
        within_pct = f"{within_pct_num}%"
        crossed_pct = f"{crossed_pct_num}%"
        result.append(['Grand Total', total_within, total_crossed, grand_total, within_pct, crossed_pct])
        
        final_report.extend(result)
    else:
        final_report.append(['Product OR PS Ticket column not found'])
    
    final_report.append([''])
    
    # 9. Ticket Group wise report
    final_report.append(['TICKET GROUP WISE REPORT'])
    if 'Ticket Group' in open_tickets.columns:
        tg_report = open_tickets.groupby(['Ticket Group', 'SLA_Status']).size().unstack(fill_value=0)
        
        result = []
        result.append(['Ticket Group', 'Within SLA', 'Crossed SLA', 'Grand Total', 'Within SLA%', 'Crossed SLA%'])
        
        total_within = 0
        total_crossed = 0
        data_rows = []
        
        for group in tg_report.index:
            within_sla = tg_report.loc[group, 'Within SLA'] if 'Within SLA' in tg_report.columns else 0
            crossed_sla = tg_report.loc[group, 'Crossed SLA'] if 'Crossed SLA' in tg_report.columns else 0
            total = within_sla + crossed_sla
            
            within_pct_num = round(within_sla * 100 / total) if total > 0 else 0
            crossed_pct_num = round(crossed_sla * 100 / total) if total > 0 else 0
            # Ensure percentages add up to 100%
            if total > 0 and within_pct_num + crossed_pct_num != 100:
                within_pct_num = 100 - crossed_pct_num
            within_pct = f"{within_pct_num}%"
            crossed_pct = f"{crossed_pct_num}%"
            
            data_rows.append([group, within_sla, crossed_sla, total, within_pct, crossed_pct, crossed_pct_num])
            
            total_within += within_sla
            total_crossed += crossed_sla
        
        # Sort by crossed SLA percentage (descending)
        data_rows.sort(key=lambda x: x[6], reverse=True)
        
        # Add sorted rows (without the sort key)
        for row in data_rows:
            result.append(row[:6])
        
        # Grand total
        grand_total = total_within + total_crossed
        within_pct_num = round(total_within * 100 / grand_total) if grand_total > 0 else 0
        crossed_pct_num = round(total_crossed * 100 / grand_total) if grand_total > 0 else 0
        if grand_total > 0 and within_pct_num + crossed_pct_num != 100:
            within_pct_num = 100 - crossed_pct_num
        within_pct = f"{within_pct_num}%"
        crossed_pct = f"{crossed_pct_num}%"
        result.append(['Grand Total', total_within, total_crossed, grand_total, within_pct, crossed_pct])
        
        final_report.extend(result)
    else:
        final_report.append(['Ticket Group column not found'])
    
    final_report.append([''])
    
    # 10. Priority wise report
    final_report.append(['PRIORITY WISE REPORT'])
    if 'Priority (Ticket)' in open_tickets.columns:
        priority_report = open_tickets.groupby(['Priority (Ticket)', 'SLA_Status']).size().unstack(fill_value=0)
        
        result = []
        result.append(['Priority (Ticket)', 'Within SLA', 'Crossed SLA', 'Grand Total', 'Within SLA%', 'Crossed SLA%'])
        
        total_within = 0
        total_crossed = 0
        data_rows = []
        
        for priority in priority_report.index:
            within_sla = priority_report.loc[priority, 'Within SLA'] if 'Within SLA' in priority_report.columns else 0
            crossed_sla = priority_report.loc[priority, 'Crossed SLA'] if 'Crossed SLA' in priority_report.columns else 0
            total = within_sla + crossed_sla
            
            within_pct_num = round(within_sla * 100 / total) if total > 0 else 0
            crossed_pct_num = round(crossed_sla * 100 / total) if total > 0 else 0
            # Ensure percentages add up to 100%
            if total > 0 and within_pct_num + crossed_pct_num != 100:
                within_pct_num = 100 - crossed_pct_num
            within_pct = f"{within_pct_num}%"
            crossed_pct = f"{crossed_pct_num}%"
            
            data_rows.append([priority, within_sla, crossed_sla, total, within_pct, crossed_pct, crossed_pct_num])
            
            total_within += within_sla
            total_crossed += crossed_sla
        
        # Sort by crossed SLA percentage (descending)
        data_rows.sort(key=lambda x: x[6], reverse=True)
        
        # Add sorted rows (without the sort key)
        for row in data_rows:
            result.append(row[:6])
        
        # Grand total
        grand_total = total_within + total_crossed
        within_pct_num = round(total_within * 100 / grand_total) if grand_total > 0 else 0
        crossed_pct_num = round(total_crossed * 100 / grand_total) if grand_total > 0 else 0
        if grand_total > 0 and within_pct_num + crossed_pct_num != 100:
            within_pct_num = 100 - crossed_pct_num
        within_pct = f"{within_pct_num}%"
        crossed_pct = f"{crossed_pct_num}%"
        result.append(['Grand Total', total_within, total_crossed, grand_total, within_pct, crossed_pct])
        
        final_report.extend(result)
    else:
        final_report.append(['Priority (Ticket) column not found'])
    
    final_report.append([''])
    
    # 11. Total Open Tickets - Detailed list
    final_report.append(['TOTAL OPEN TICKETS'])
    
    # Add header with all requested columns
    header = ['Gitlab Link', 'Select Engineer', 'Module Lead', 'Ticket Group', 'Product OR PS Ticket', 'Program Name', 'Department Name', 'Subject']
    final_report.append(header)
    
    # Add all open tickets data
    for _, ticket in open_tickets.iterrows():
        row = [
            ticket.get('Gitlab Link', ''),
            ticket.get('Select Engineer', ''),
            ticket.get('Module Lead', ''),
            ticket.get('Ticket Group', ''),
            ticket.get('Product OR PS Ticket', ''),
            ticket.get('Program Name', ''),
            ticket.get('Department Name', ''),
            ticket.get('Subject', '')
        ]
        final_report.append(row)
    
    return pd.DataFrame(final_report)

def process_recurring_issues_mis(df):
    """Process Advanced Recurring Issues MIS with intelligent pattern matching and comprehensive analysis"""
    import re
    from difflib import SequenceMatcher
    import datetime
    
    # Check for required columns
    resolution_cols = ['Resolution', 'Solution', 'Fix', 'Root Cause', 'Closure Comments', 'Subject']
    resolution_col = None
    for col in resolution_cols:
        if col in df.columns:
            resolution_col = col
            break
    
    if resolution_col is None:
        return pd.DataFrame({'Error': ['Resolution/Subject column not found. Expected: Resolution, Solution, Fix, Root Cause, Closure Comments, or Subject']})
    
    # Enhanced text normalization with better pattern recognition
    def normalize_text(text):
        if pd.isna(text):
            return ''
        text = str(text).lower().strip()
        
        # Remove common variable data but preserve meaningful patterns
        text = re.sub(r'\b[a-z]{2,4}-\d+\b', '[ticket_id]', text)  # Ticket IDs
        text = re.sub(r'\b[a-z0-9]{8,20}\b', '[reference]', text)  # Reference numbers
        text = re.sub(r'\b\d{10,12}\b', '[phone]', text)  # Phone numbers
        text = re.sub(r'\b[a-z0-9._%+-]+@[a-z0-9.-]+\.[a-z]{2,}\b', '[email]', text)  # Emails
        text = re.sub(r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b', '[date]', text)  # Dates
        text = re.sub(r'[₹$€£]\s*\d+[.,]?\d*', '[amount]', text)  # Currency amounts
        text = re.sub(r'\b\d{4,}\b', '[number]', text)  # Large numbers
        text = re.sub(r'version\s*\d+\.\d+', 'version [x.x]', text)  # Version numbers
        text = re.sub(r'\s+', ' ', text).strip()
        
        # Extract key error patterns
        error_patterns = [
            'error', 'failed', 'timeout', 'connection', 'unable', 'cannot', 
            'invalid', 'missing', 'not found', 'access denied', 'permission',
            'server', 'database', 'network', 'api', 'sync', 'login', 'password'
        ]
        
        # Preserve important technical terms
        for pattern in error_patterns:
            if pattern in text:
                text = f"{pattern} {text}"
                break
                
        return text[:200]  # Limit length for better matching
    
    # Enhanced similarity function with weighted scoring
    def enhanced_similarity(a, b):
        # Basic sequence similarity
        seq_sim = SequenceMatcher(None, a, b).ratio()
        
        # Word-based similarity for better semantic matching
        words_a = set(a.split())
        words_b = set(b.split())
        
        if len(words_a) == 0 or len(words_b) == 0:
            return seq_sim
            
        word_sim = len(words_a.intersection(words_b)) / len(words_a.union(words_b))
        
        # Weighted combination
        return (seq_sim * 0.6) + (word_sim * 0.4)
    
    # Advanced clustering with multiple similarity thresholds
    def create_advanced_clusters(df_clean):
        clusters = []
        processed = set()
        
        # Sort by normalized text length for better clustering
        df_sorted = df_clean.sort_values('text_length', ascending=False)
        
        for idx, row in df_sorted.iterrows():
            if idx in processed:
                continue
            
            # Create new cluster
            cluster = {
                'tickets': [idx], 
                'pattern': row['normalized_text'],
                'original_subject': row.get('Subject', ''),
                'category': row.get('subcategory', '')
            }
            processed.add(idx)
            
            # Find similar tickets with adaptive threshold
            base_threshold = 0.65  # Lower threshold for better recall
            
            for idx2, row2 in df_sorted.iterrows():
                if idx2 in processed:
                    continue
                
                # Calculate similarity
                sim_score = enhanced_similarity(row['normalized_text'], row2['normalized_text'])
                
                # Adaptive threshold based on text length
                adaptive_threshold = base_threshold
                if len(row['normalized_text']) < 50:
                    adaptive_threshold = 0.75  # Higher threshold for short texts
                
                if sim_score >= adaptive_threshold:
                    cluster['tickets'].append(idx2)
                    processed.add(idx2)
            
            # Only keep clusters with 2+ tickets
            if len(cluster['tickets']) >= 2:
                clusters.append(cluster)
        
        return clusters
    
    # Prepare enhanced data
    df_clean = df.copy()
    
    # Find best available columns for analysis
    subcategory_col = None
    for col in ['Ticket Sub Category', 'Request Sub Category', 'Category Of Issue', 'Category Type', 'Subject']:
        if col in df.columns:
            subcategory_col = col
            break
    
    subject_col = 'Subject' if 'Subject' in df.columns else None
    
    # Create comprehensive text for analysis
    df_clean['subcategory'] = df_clean[subcategory_col].fillna('') if subcategory_col else ''
    df_clean['resolution'] = df_clean[resolution_col].fillna('')
    df_clean['subject'] = df_clean[subject_col].fillna('') if subject_col else ''
    
    # Combine multiple fields for better pattern recognition
    df_clean['combined_text'] = (
        df_clean['subject'].astype(str) + ' ' + 
        df_clean['subcategory'].astype(str) + ' ' + 
        df_clean['resolution'].astype(str)
    )
    
    df_clean['normalized_text'] = df_clean['combined_text'].apply(normalize_text)
    df_clean['text_length'] = df_clean['normalized_text'].str.len()
    
    # Create advanced clusters
    clusters = create_advanced_clusters(df_clean)
    
    final_report = []
    
    # 1. EXECUTIVE SUMMARY
    final_report.append(['RECURRING ISSUES EXECUTIVE SUMMARY'])
    final_report.append([''])
    
    total_tickets = len(df_clean)
    recurring_tickets = sum(len(cluster['tickets']) for cluster in clusters)
    recurring_percentage = round((recurring_tickets / total_tickets) * 100, 1) if total_tickets > 0 else 0
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Tickets Analyzed', total_tickets],
        ['Recurring Issue Patterns Found', len(clusters)],
        ['Tickets in Recurring Patterns', recurring_tickets],
        ['Recurring Issues Percentage', f'{recurring_percentage}%'],
        ['High Impact Issues (5+ occurrences)', len([c for c in clusters if len(c['tickets']) >= 5])],
        ['Critical Issues (10+ occurrences)', len([c for c in clusters if len(c['tickets']) >= 10])]
    ]
    
    final_report.extend(summary_data)
    final_report.append([''])
    final_report.append([''])
    
    # 2. TOP RECURRING ISSUES ANALYSIS
    final_report.append(['TOP RECURRING ISSUES ANALYSIS'])
    final_report.append([''])
    
    if clusters:
        result = []
        result.append(['Rank', 'Issue Pattern', 'Occurrences', 'Programs Affected', 'Engineers Involved', 'Resolution Rate', 'Avg Reopens', 'First Occurrence', 'Last Occurrence', 'Impact Level', 'Recommended Action'])
        
        # Sort clusters by impact (occurrences * reopens)
        def calculate_impact(cluster):
            cluster_tickets = df_clean.loc[cluster['tickets']]
            occurrences = len(cluster_tickets)
            if 'Number of Reopen' in df.columns:
                reopens = pd.to_numeric(cluster_tickets['Number of Reopen'], errors='coerce').fillna(0).sum()
                return occurrences + (reopens * 2)  # Weight reopens more heavily
            return occurrences
        
        sorted_clusters = sorted(clusters, key=calculate_impact, reverse=True)[:25]
        
        for rank, cluster in enumerate(sorted_clusters, 1):
            cluster_tickets = df_clean.loc[cluster['tickets']]
            occurrences = len(cluster_tickets)
            
            # Programs affected
            programs_affected = cluster_tickets['Program Name'].nunique() if 'Program Name' in df.columns else 'Unknown'
            
            # Engineers involved
            if 'Select Engineer' in df.columns:
                engineers = cluster_tickets['Select Engineer'].dropna().nunique()
            else:
                engineers = 'Unknown'
            
            # Resolution rate
            if 'Status (Ticket)' in df.columns:
                closed_count = len(cluster_tickets[cluster_tickets['Status (Ticket)'] == 'Closed'])
                resolution_rate = f"{round((closed_count/occurrences)*100)}%"
            else:
                resolution_rate = 'Unknown'
            
            # Average reopens
            if 'Number of Reopen' in df.columns:
                avg_reopens = round(pd.to_numeric(cluster_tickets['Number of Reopen'], errors='coerce').fillna(0).mean(), 1)
            else:
                avg_reopens = 0
            
            # Date analysis
            date_col = None
            for col in ['Created Time (Ticket)', 'Created Tim']:
                if col in df.columns:
                    date_col = col
                    break
            
            if date_col:
                dates = pd.to_datetime(cluster_tickets[date_col], errors='coerce').dropna()
                if not dates.empty:
                    first_date = dates.min().strftime('%Y-%m-%d')
                    last_date = dates.max().strftime('%Y-%m-%d')
                else:
                    first_date = last_date = 'Unknown'
            else:
                first_date = last_date = 'Unknown'
            
            # Impact level
            if occurrences >= 10:
                impact_level = 'Critical'
            elif occurrences >= 5:
                impact_level = 'High'
            elif occurrences >= 3:
                impact_level = 'Medium'
            else:
                impact_level = 'Low'
            
            # Recommended action
            pattern_lower = cluster['pattern'].lower()
            if avg_reopens > 1:
                action = 'Root Cause Analysis & Process Review'
            elif programs_affected > 3:
                action = 'System-wide Investigation'
            elif 'error' in pattern_lower or 'failed' in pattern_lower:
                action = 'Technical Fix Required'
            elif 'timeout' in pattern_lower or 'connection' in pattern_lower:
                action = 'Infrastructure Review'
            else:
                action = 'Process Improvement'
            
            # Truncate pattern for display
            display_pattern = cluster['pattern'][:80] + '...' if len(cluster['pattern']) > 80 else cluster['pattern']
            
            result.append([
                rank,
                display_pattern,
                occurrences,
                programs_affected,
                engineers,
                resolution_rate,
                avg_reopens,
                first_date,
                last_date,
                impact_level,
                action
            ])
        
        final_report.extend(result)
    else:
        final_report.append(['No recurring patterns found with current similarity threshold'])
    
    final_report.append([''])
    final_report.append([''])
    
    # 3. TICKET SUB CATEGORY-WISE RECURRING ISSUES
    final_report.append(['TICKET SUB CATEGORY-WISE RECURRING ISSUES BREAKDOWN'])
    final_report.append([''])
    
    if clusters and subcategory_col:
        category_analysis = {}
        for cluster in clusters:
            cluster_tickets = df_clean.loc[cluster['tickets']]
            for category in cluster_tickets[subcategory_col].dropna().unique():
                if category not in category_analysis:
                    category_analysis[category] = {'clusters': 0, 'tickets': 0}
                category_analysis[category]['clusters'] += 1
                category_analysis[category]['tickets'] += len(cluster_tickets[cluster_tickets[subcategory_col] == category])
        
        result = []
        result.append(['Ticket Sub Category', 'Recurring Patterns', 'Total Recurring Tickets', 'Avg Tickets per Pattern'])
        
        for category, data in sorted(category_analysis.items(), key=lambda x: x[1]['tickets'], reverse=True):
            avg_tickets = round(data['tickets'] / data['clusters'], 1) if data['clusters'] > 0 else 0
            result.append([category, data['clusters'], data['tickets'], avg_tickets])
        
        final_report.extend(result)
    else:
        final_report.append(['Ticket Sub Category information not available'])
    
    final_report.append([''])
    final_report.append([''])
    
    # 4. ENGINEER PERFORMANCE ON RECURRING ISSUES
    final_report.append(['ENGINEER PERFORMANCE ON RECURRING ISSUES'])
    final_report.append([''])
    
    if 'Select Engineer' in df.columns and clusters:
        result = []
        result.append(['Engineer', 'Recurring Issues Handled', 'Unique Patterns', 'Resolution Rate', 'Avg Reopens', 'Performance Score', 'Focus Area'])
        
        # Get all tickets in clusters
        cluster_ticket_ids = set()
        for cluster in clusters:
            cluster_ticket_ids.update(cluster['tickets'])
        
        recurring_tickets = df_clean.loc[list(cluster_ticket_ids)]
        
        engineer_data = []
        for engineer in recurring_tickets['Select Engineer'].dropna().unique():
            engineer_tickets = recurring_tickets[recurring_tickets['Select Engineer'] == engineer]
            
            issues_handled = len(engineer_tickets)
            
            # Count unique patterns this engineer worked on
            engineer_patterns = set()
            for cluster in clusters:
                cluster_tickets = df_clean.loc[cluster['tickets']]
                if engineer in cluster_tickets['Select Engineer'].values:
                    engineer_patterns.add(cluster['pattern'])
            unique_patterns = len(engineer_patterns)
            
            # Resolution rate
            if 'Status (Ticket)' in df.columns:
                resolved = len(engineer_tickets[engineer_tickets['Status (Ticket)'] == 'Closed'])
                resolution_rate = round((resolved/issues_handled)*100, 1) if issues_handled > 0 else 0
            else:
                resolution_rate = 0
            
            # Average reopens
            if 'Number of Reopen' in df.columns:
                avg_reopens = round(pd.to_numeric(engineer_tickets['Number of Reopen'], errors='coerce').fillna(0).mean(), 1)
            else:
                avg_reopens = 0
            
            # Performance score (weighted combination)
            performance_score = round(
                (resolution_rate * 0.4) + 
                ((100 - min(avg_reopens * 10, 100)) * 0.3) + 
                (min(unique_patterns * 5, 50) * 0.3), 1
            )
            
            # Focus area recommendation
            if resolution_rate < 70:
                focus_area = 'Resolution Efficiency'
            elif avg_reopens > 1.5:
                focus_area = 'Quality Improvement'
            elif unique_patterns < 3:
                focus_area = 'Knowledge Expansion'
            else:
                focus_area = 'Mentoring Others'
            
            engineer_data.append([
                engineer, 
                issues_handled, 
                unique_patterns, 
                f"{resolution_rate}%", 
                avg_reopens, 
                performance_score,
                focus_area
            ])
        
        # Sort by performance score (descending)
        engineer_data.sort(key=lambda x: x[5], reverse=True)
        result.extend(engineer_data)
        final_report.extend(result)
    else:
        final_report.append(['Engineer data not available'])
    
    final_report.append([''])
    final_report.append([''])
    
    # 5. MONTHLY TREND ANALYSIS
    final_report.append(['RECURRING ISSUES MONTHLY TREND ANALYSIS'])
    final_report.append([''])
    
    date_col = None
    for col in ['Created Time (Ticket)', 'Created Tim']:
        if col in df.columns:
            date_col = col
            break
    
    if clusters and date_col:
        result = []
        result.append(['Month', 'New Patterns', 'Total Occurrences', 'Critical Issues', 'Resolution Rate', 'Trend'])
        
        today = datetime.datetime.now()
        monthly_data = []
        
        for i in range(6, 0, -1):  # Last 6 months
            month_start = (today.replace(day=1) - datetime.timedelta(days=30*i)).replace(day=1)
            month_end = (month_start + datetime.timedelta(days=32)).replace(day=1) - datetime.timedelta(days=1)
            period = month_start.strftime('%Y-%m')
            
            # Count patterns and occurrences in this period
            period_patterns = 0
            period_occurrences = 0
            period_critical = 0
            period_resolved = 0
            period_total = 0
            
            for cluster in clusters:
                cluster_tickets = df_clean.loc[cluster['tickets']]
                dates = pd.to_datetime(cluster_tickets[date_col], errors='coerce')
                period_tickets = cluster_tickets[(dates >= month_start) & (dates <= month_end)]
                
                if not period_tickets.empty:
                    period_patterns += 1
                    period_occurrences += len(period_tickets)
                    
                    if len(cluster['tickets']) >= 5:  # Critical threshold
                        period_critical += 1
                    
                    # Resolution rate for this period
                    if 'Status (Ticket)' in df.columns:
                        resolved = len(period_tickets[period_tickets['Status (Ticket)'] == 'Closed'])
                        period_resolved += resolved
                        period_total += len(period_tickets)
            
            resolution_rate = round((period_resolved / period_total) * 100, 1) if period_total > 0 else 0
            
            monthly_data.append({
                'month': period,
                'patterns': period_patterns,
                'occurrences': period_occurrences,
                'critical': period_critical,
                'resolution_rate': resolution_rate
            })
        
        # Calculate trends
        for i, data in enumerate(monthly_data):
            if i == 0:
                trend = 'Baseline'
            else:
                prev_occurrences = monthly_data[i-1]['occurrences']
                curr_occurrences = data['occurrences']
                
                if curr_occurrences > prev_occurrences * 1.1:
                    trend = 'Increasing ↗'
                elif curr_occurrences < prev_occurrences * 0.9:
                    trend = 'Decreasing ↘'
                else:
                    trend = 'Stable →'
            
            result.append([
                data['month'],
                data['patterns'],
                data['occurrences'],
                data['critical'],
                f"{data['resolution_rate']}%",
                trend
            ])
        
        final_report.extend(result)
    else:
        final_report.append(['Date information not available for trend analysis'])
    

    
    return pd.DataFrame(final_report)

def generate_client_closed_report(closed_tickets, program):
    """Generate closed tickets report for a specific program"""
    result = []
    result.append(['Client Name', 'Closed Tickets Within SLA', 'Closed Tickets Crossed SLA', 'Total Closed Tickets', 'Within SLA%', 'Crossed SLA%'])
    
    # Use program name as client name and aggregate all tickets
    within_sla = len(closed_tickets[closed_tickets['SLA_Status'] == 'Within SLA'])
    crossed_sla = len(closed_tickets[closed_tickets['SLA_Status'] == 'Crossed SLA'])
    total = within_sla + crossed_sla
    
    within_pct_num = round(within_sla * 100 / total) if total > 0 else 0
    crossed_pct_num = round(crossed_sla * 100 / total) if total > 0 else 0
    if total > 0 and within_pct_num + crossed_pct_num != 100:
        within_pct_num = 100 - crossed_pct_num
    within_pct = f"{within_pct_num}%"
    crossed_pct = f"{crossed_pct_num}%"
    
    result.append([program, within_sla, crossed_sla, total, within_pct, crossed_pct])
    
    return pd.DataFrame(result[1:], columns=result[0])

def generate_client_open_report(open_tickets, program):
    """Generate open tickets report for a specific program"""
    result = []
    result.append(['Client Name', 'Open tickets within SLA', 'Open Tickets Crossed SLA', 'Total Open Tickets', 'Within SLA%', 'Crossed SLA%'])
    
    # Use program name as client name and aggregate all tickets
    within_sla = len(open_tickets[open_tickets['SLA_Status'] == 'Within SLA'])
    crossed_sla = len(open_tickets[open_tickets['SLA_Status'] == 'Crossed SLA'])
    total = within_sla + crossed_sla
    
    within_pct_num = round(within_sla * 100 / total) if total > 0 else 0
    crossed_pct_num = round(crossed_sla * 100 / total) if total > 0 else 0
    if total > 0 and within_pct_num + crossed_pct_num != 100:
        within_pct_num = 100 - crossed_pct_num
    within_pct = f"{within_pct_num}%"
    crossed_pct = f"{crossed_pct_num}%"
    
    result.append([program, within_sla, crossed_sla, total, within_pct, crossed_pct])
    
    return pd.DataFrame(result[1:], columns=result[0])

def generate_client_request_report(program_df, program):
    """Generate request tickets report for a specific program"""
    result = []
    result.append(['Client Name', 'Request Closed', 'Request Open', 'Grand Total'])
    
    # Filter request tickets based on Classifications column
    if 'Classifications' in program_df.columns:
        request_tickets = program_df[program_df['Classifications'].str.lower().str.contains('request', na=False)]
    else:
        # If no Classifications column, assume all are request tickets
        request_tickets = program_df
    
    # Use program name as client name and aggregate all request tickets
    closed_count = len(request_tickets[request_tickets['Status (Ticket)'] == 'Closed'])
    open_statuses = ['Assigned to Engineer!', 'Reopened', 'Waiting Information From user - 1', 'Waiting Information From user - 2', 'Waiting Information From user - 3']
    open_count = len(request_tickets[request_tickets['Status (Ticket)'].isin(open_statuses)])
    total = closed_count + open_count
    
    result.append([program, closed_count, open_count, total])
    
    return pd.DataFrame(result[1:], columns=result[0])

def process_open_ticket_mis(df):
    """
    Process Open Ticket MIS to generate Module Lead, Client, and Engineer wise reports
    """
    # Check if 'Status (Ticket)' column exists
    if 'Status (Ticket)' not in df.columns:
        return pd.DataFrame({'Error': ['Status (Ticket) column not found']})
    
    # Filter open tickets with all specified statuses
    open_statuses = [
        'Assigned to Engineer!',
        'Reopened',
        'Waiting Information From user - 1',
        'Waiting Information From user - 2', 
        'Waiting Information From user - 3'
    ]
    open_tickets = df[df['Status (Ticket)'].isin(open_statuses)].copy()
    
    # Only exclude tickets that are in waiting status AND classified as 'request open'
    waiting_statuses = [
        'Waiting Information From user - 1',
        'Waiting Information From user - 2',
        'Waiting Information From user - 3'
    ]
    
    if 'Classifications' in open_tickets.columns:
        # Remove tickets that are both in waiting status AND classified as request open
        exclude_condition = (
            open_tickets['Status (Ticket)'].isin(waiting_statuses) & 
            (open_tickets['Classifications'].str.lower().str.contains('request open', na=False))
        )
        open_tickets = open_tickets[~exclude_condition]
    
    if open_tickets.empty:
        return pd.DataFrame({'Error': ['No open tickets found']})
    
    # Calculate SLA status based on GitLab due date
    import datetime
    today_date = datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    def calculate_sla_status(row):
        if 'Gitlab Due date' in open_tickets.columns and pd.notna(row.get('Gitlab Due date')):
            try:
                gitlab_due = pd.to_datetime(row['Gitlab Due date'], errors='coerce')
                if pd.notna(gitlab_due):
                    return 'Crossed SLA' if today_date.date() > gitlab_due.date() else 'Within SLA'
            except:
                pass
        # Fallback to Is Overdue if Gitlab due date not available
        return 'Crossed SLA' if row.get('Is Overdue') == True else 'Within SLA'
    
    open_tickets['SLA_Status'] = open_tickets.apply(calculate_sla_status, axis=1)
    
    # Sort by creation date (ascending) to show oldest tickets first
    if 'Created Time (Ticket)' in open_tickets.columns:
        open_tickets['Created_Date_Sort'] = pd.to_datetime(open_tickets['Created Time (Ticket)'], errors='coerce')
        open_tickets = open_tickets.sort_values('Created_Date_Sort', ascending=True)
        open_tickets = open_tickets.drop('Created_Date_Sort', axis=1)
    
    # Generate all three reports
    module_lead_report = generate_module_lead_report(open_tickets)
    client_report = generate_client_report(open_tickets)
    engineer_report = generate_engineer_report(open_tickets)
    
    # Create structured report with proper headers
    final_report = []
    
    # Module Lead Report
    final_report.append(['MODULE LEAD WISE REPORT'])
    final_report.extend(module_lead_report.values.tolist())
    final_report.append([''])
    
    # Client Report  
    final_report.append(['CLIENT WISE REPORT'])
    final_report.extend(client_report.values.tolist())
    final_report.append([''])
    
    # Engineer Report
    final_report.append(['ENGINEER WISE REPORT'])
    final_report.extend(engineer_report.values.tolist())
    
    return pd.DataFrame(final_report)

def generate_module_lead_report(open_tickets):
    """Generate Module Lead wise report"""
    report = open_tickets.groupby(['Module Lead', 'SLA_Status']).size().unstack(fill_value=0)
    
    result = []
    result.append(['Module Lead', 'Within SLA', 'Crossed SLA', 'Grand Total', 'Within SLA%', 'Crossed SLA%'])
    
    total_within = 0
    total_crossed = 0
    data_rows = []
    
    for module_lead in report.index:
        within_sla = report.loc[module_lead, 'Within SLA'] if 'Within SLA' in report.columns else 0
        crossed_sla = report.loc[module_lead, 'Crossed SLA'] if 'Crossed SLA' in report.columns else 0
        total = within_sla + crossed_sla
        
        within_pct_num = round(within_sla * 100 / total) if total > 0 else 0
        crossed_pct_num = round(crossed_sla * 100 / total) if total > 0 else 0
        # Ensure percentages add up to 100%
        if total > 0 and within_pct_num + crossed_pct_num != 100:
            within_pct_num = 100 - crossed_pct_num
        within_pct = f"{within_pct_num}%"
        crossed_pct = f"{crossed_pct_num}%"
        
        data_rows.append([module_lead, within_sla, crossed_sla, total, within_pct, crossed_pct, crossed_pct_num])
        
        total_within += within_sla
        total_crossed += crossed_sla
    
    # Sort by crossed SLA percentage (descending)
    data_rows.sort(key=lambda x: x[6], reverse=True)
    
    # Add sorted rows (without the sort key)
    for row in data_rows:
        result.append(row[:6])
    
    # Grand total
    grand_total = total_within + total_crossed
    within_pct_num = round(total_within * 100 / grand_total) if grand_total > 0 else 0
    crossed_pct_num = round(total_crossed * 100 / grand_total) if grand_total > 0 else 0
    if grand_total > 0 and within_pct_num + crossed_pct_num != 100:
        within_pct_num = 100 - crossed_pct_num
    within_pct = f"{within_pct_num}%"
    crossed_pct = f"{crossed_pct_num}%"
    result.append(['Grand Total', total_within, total_crossed, grand_total, within_pct, crossed_pct])
    
    return pd.DataFrame(result[1:], columns=result[0])

def generate_client_report(open_tickets):
    """Generate Client wise report"""
    # Use Program Name as Client Name
    report = open_tickets.groupby(['Program Name', 'SLA_Status']).size().unstack(fill_value=0)
    
    result = []
    result.append(['Client Name', 'Within SLA', 'Crossed SLA', 'Grand Total', 'Within SLA%', 'Crossed SLA%'])
    
    total_within = 0
    total_crossed = 0
    data_rows = []
    
    for client in report.index:
        within_sla = report.loc[client, 'Within SLA'] if 'Within SLA' in report.columns else 0
        crossed_sla = report.loc[client, 'Crossed SLA'] if 'Crossed SLA' in report.columns else 0
        total = within_sla + crossed_sla
        
        within_pct_num = round(within_sla * 100 / total) if total > 0 else 0
        crossed_pct_num = round(crossed_sla * 100 / total) if total > 0 else 0
        # Ensure percentages add up to 100%
        if total > 0 and within_pct_num + crossed_pct_num != 100:
            within_pct_num = 100 - crossed_pct_num
        within_pct = f"{within_pct_num}%"
        crossed_pct = f"{crossed_pct_num}%"
        
        data_rows.append([client, within_sla, crossed_sla, total, within_pct, crossed_pct, crossed_pct_num])
        
        total_within += within_sla
        total_crossed += crossed_sla
    
    # Sort by crossed SLA percentage (descending)
    data_rows.sort(key=lambda x: x[6], reverse=True)
    
    # Add sorted rows (without the sort key)
    for row in data_rows:
        result.append(row[:6])
    
    # Grand total
    grand_total = total_within + total_crossed
    within_pct_num = round(total_within * 100 / grand_total) if grand_total > 0 else 0
    crossed_pct_num = round(total_crossed * 100 / grand_total) if grand_total > 0 else 0
    if grand_total > 0 and within_pct_num + crossed_pct_num != 100:
        within_pct_num = 100 - crossed_pct_num
    within_pct = f"{within_pct_num}%"
    crossed_pct = f"{crossed_pct_num}%"
    result.append(['Grand Total', total_within, total_crossed, grand_total, within_pct, crossed_pct])
    
    return pd.DataFrame(result[1:], columns=result[0])

def generate_engineer_report(open_tickets):
    """Generate Engineer wise report"""
    report = open_tickets.groupby(['Select Engineer', 'SLA_Status']).size().unstack(fill_value=0)
    
    result = []
    result.append(['Engineer', 'Within SLA', 'Crossed SLA', 'Grand Total', 'Within SLA%', 'Crossed SLA%'])
    
    total_within = 0
    total_crossed = 0
    data_rows = []
    
    for engineer in report.index:
        within_sla = report.loc[engineer, 'Within SLA'] if 'Within SLA' in report.columns else 0
        crossed_sla = report.loc[engineer, 'Crossed SLA'] if 'Crossed SLA' in report.columns else 0
        total = within_sla + crossed_sla
        
        within_pct_num = round(within_sla * 100 / total) if total > 0 else 0
        crossed_pct_num = round(crossed_sla * 100 / total) if total > 0 else 0
        # Ensure percentages add up to 100%
        if total > 0 and within_pct_num + crossed_pct_num != 100:
            within_pct_num = 100 - crossed_pct_num
        within_pct = f"{within_pct_num}%"
        crossed_pct = f"{crossed_pct_num}%"
        
        data_rows.append([engineer, within_sla, crossed_sla, total, within_pct, crossed_pct, crossed_pct_num])
        
        total_within += within_sla
        total_crossed += crossed_sla
    
    # Sort by crossed SLA percentage (descending)
    data_rows.sort(key=lambda x: x[6], reverse=True)
    
    # Add sorted rows (without the sort key)
    for row in data_rows:
        result.append(row[:6])
    
    # Grand total
    grand_total = total_within + total_crossed
    within_pct_num = round(total_within * 100 / grand_total) if grand_total > 0 else 0
    crossed_pct_num = round(total_crossed * 100 / grand_total) if grand_total > 0 else 0
    if grand_total > 0 and within_pct_num + crossed_pct_num != 100:
        within_pct_num = 100 - crossed_pct_num
    within_pct = f"{within_pct_num}%"
    crossed_pct = f"{crossed_pct_num}%"
    result.append(['Grand Total', total_within, total_crossed, grand_total, within_pct, crossed_pct])
    
    return pd.DataFrame(result[1:], columns=result[0])

def process_request_ticket_open_mis(df):
    """
    Process Request Ticket Open MIS:
    1. Add today's date as datetime
    2. Calculate days difference between today and L1-Due Date (not GitLab due date)
    3. Generate both raw data sheet and MIS summary sheet
    """
    import datetime
    
    # Check required columns
    required_cols = ['Created Time (Ticket)', 'Created Tim']
    if not any(col in df.columns for col in required_cols):
        return pd.DataFrame({'Error': ['Created Time column not found']})
    
    # Create a copy of the dataframe
    result_df = df.copy()
    
    # Add today's date as datetime (matching expected format)
    today_date = datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    result_df['Todays Date'] = today_date
    
    # Calculate days difference using today's date - created date
    def calculate_days_diff(created_date):
        try:
            if pd.isna(created_date) or created_date == '':
                return 0
            
            # Parse Created Date
            created_dt = pd.to_datetime(created_date, errors='coerce')
            if pd.isna(created_dt):
                return 0
            
            # Calculate difference (today - created date)
            diff = (today_date - created_dt).days
            return max(0, diff)
        except:
            return 0
    
    # Use Created Tim column for calculation
    created_col = 'Created Tim' if 'Created Tim' in result_df.columns else 'Created Time (Ticket)'
    result_df['No of crossed days'] = result_df[created_col].apply(calculate_days_diff)
    
    # Match exact column order from expected output
    expected_columns = [
        'Ticket Id', 'Status (Ticket)', 'Ticket Owner', 'Created Tim', 
        'Due Date', 'Ticket Closed Time', 'Email (Contact)', 'Category Type',
        'Priority (Ticket)', 'Severity Classification', 'Channel',
        'Total Time Spent', 'Crossed Due Date', 'L1-Due Date', 
        'Request Sub Category', 'Contact name', 'Category Of Issue', 
        'Gitlab Due date', 'Gitlab Link', 'Number of Reopen',
        'Is Overdue', 'Support Plan Category', 'Classifications', 'Ticket Group',
        'Solutions Engineer', 'Select Engineer', 'Module Lead', 'Program Name',
        'Department Name', 'Product OR PS Ticket', 'Subject', 
        'Todays Date', 'No of crossed days'
    ]
    
    # Map column names from raw data to expected output format
    column_mapping = {
        'Created Time (Ticket)': 'Created Tim',  # Handle truncated column name
        'Account Name': 'Contact name'  # Map if Account Name exists instead of Contact name
    }
    
    # Rename columns if needed
    for old_name, new_name in column_mapping.items():
        if old_name in result_df.columns and new_name not in result_df.columns:
            result_df = result_df.rename(columns={old_name: new_name})
    
    # Keep only columns that exist and match expected order
    available_columns = [col for col in expected_columns if col in result_df.columns]
    result_df = result_df[available_columns]
    
    # Generate MIS summary
    mis_summary = generate_request_ticket_mis_summary(result_df)
    
    return {'raw_data': result_df, 'mis_summary': mis_summary}

def generate_request_ticket_mis_summary(df):
    """
    Generate MIS summary with Program Names/Engineers in rows and No of crossed days as columns
    """
    final_report = []
    
    # 1. Solutions Engineer wise report
    final_report.append(['SOLUTIONS ENGINEER WISE REPORT'])
    se_report = df.groupby(['Solutions Engineer', 'No of crossed days']).size().unstack(fill_value=0)
    se_report['Grand Total'] = se_report.sum(axis=1)
    
    # Only include columns that have non-zero values
    day_cols = sorted([col for col in se_report.columns if col != 'Grand Total' and se_report[col].sum() > 0])
    
    # Add header
    header = ['Solutions Engineer'] + [str(col) for col in day_cols] + ['Grand Total']
    final_report.append(header)
    
    # Add data rows
    for engineer in se_report.index:
        row = [engineer] + [se_report.loc[engineer, col] for col in day_cols] + [se_report.loc[engineer, 'Grand Total']]
        final_report.append(row)
    
    # Add grand total row
    grand_total_row = ['Grand Total'] + [se_report[col].sum() for col in day_cols] + [se_report['Grand Total'].sum()]
    final_report.append(grand_total_row)
    final_report.append([''])  # Empty row
    
    # 2. Program Name wise report
    final_report.append(['PROGRAM NAME WISE REPORT'])
    pn_report = df.groupby(['Program Name', 'No of crossed days']).size().unstack(fill_value=0)
    pn_report['Grand Total'] = pn_report.sum(axis=1)
    
    # Only include columns that have non-zero values
    day_cols = sorted([col for col in pn_report.columns if col != 'Grand Total' and pn_report[col].sum() > 0])
    
    # Add header
    header = ['Program Name'] + [str(col) for col in day_cols] + ['Grand Total']
    final_report.append(header)
    
    # Add data rows
    for program in pn_report.index:
        row = [program] + [pn_report.loc[program, col] for col in day_cols] + [pn_report.loc[program, 'Grand Total']]
        final_report.append(row)
    
    # Add grand total row
    grand_total_row = ['Grand Total'] + [pn_report[col].sum() for col in day_cols] + [pn_report['Grand Total'].sum()]
    final_report.append(grand_total_row)
    final_report.append([''])  # Empty row
    
    # 3. Select Engineer wise report
    final_report.append(['SELECT ENGINEER WISE REPORT'])
    eng_report = df.groupby(['Select Engineer', 'No of crossed days']).size().unstack(fill_value=0)
    eng_report['Grand Total'] = eng_report.sum(axis=1)
    
    # Only include columns that have non-zero values
    day_cols = sorted([col for col in eng_report.columns if col != 'Grand Total' and eng_report[col].sum() > 0])
    
    # Add header
    header = ['Select Engineer'] + [str(col) for col in day_cols] + ['Grand Total']
    final_report.append(header)
    
    # Add data rows
    for engineer in eng_report.index:
        row = [engineer] + [eng_report.loc[engineer, col] for col in day_cols] + [eng_report.loc[engineer, 'Grand Total']]
        final_report.append(row)
    
    # Add grand total row
    grand_total_row = ['Grand Total'] + [eng_report[col].sum() for col in day_cols] + [eng_report['Grand Total'].sum()]
    final_report.append(grand_total_row)
    
    return pd.DataFrame(final_report)

if __name__ == "__main__":
    st.set_page_config(
        page_title="MIS Support Bot",
        page_icon="🤖",
        layout="wide"
    )
    main()
